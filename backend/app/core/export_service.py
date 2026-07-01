"""
Export Service for IP Scan Results
Generates professional Excel and PDF reports
"""

from io import BytesIO
from typing import Dict, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


class ExportService:
    """Service for exporting IP scan results to Excel and PDF"""
    
    def __init__(self):
        self.severity_colors = {
            "Low": "4CAF50",
            "Medium": "FFC107",
            "High": "FF9800",
            "Critical": "F44336"
        }
    
    def export_to_excel(self, scan_data: Dict[str, Any]) -> BytesIO:
        """Export scan results to Excel with professional formatting"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self._create_executive_summary_sheet(wb, scan_data)
        self._create_network_location_sheet(wb, scan_data)
        self._create_threat_intel_sheet(wb, scan_data)
        self._create_details_sheet(wb, scan_data)
        self._create_raw_data_sheet(wb, scan_data)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_raw_data_sheet(self, wb: openpyxl.Workbook, scan_data: Dict[str, Any]):
        """Create raw JSON data sheet"""
        ws = wb.create_sheet("Raw Data")
        import json
        ws['A1'] = "Raw Scan Data (JSON)"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = json.dumps(scan_data, indent=4, default=str)
        ws.column_dimensions['A'].width = 100
        ws.row_dimensions[2].height = 400
        ws['A2'].alignment = Alignment(wrap_text=True, vertical="top")

    def _create_executive_summary_sheet(self, wb: openpyxl.Workbook, scan_data: Dict[str, Any]):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "IP Security Scan Report"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Scan info
        risk = scan_data.get("risk_assessment", {})
        geo = scan_data.get("geolocation", {})
        threat = scan_data.get("threat_intelligence", {})
        
        row = 3
        ws[f'A{row}'] = "IP Address:"
        ws[f'B{row}'] = scan_data.get("ip", "N/A")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Scan Date:"
        ws[f'B{row}'] = scan_data.get("scan_timestamp", datetime.utcnow().isoformat())
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 2
        ws[f'A{row}'] = "RISK ASSESSMENT"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "Risk Score:"
        ws[f'B{row}'] = f"{risk.get('score', 0)}/100"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Severity:"
        severity = risk.get('severity', 'Low')
        ws[f'B{row}'] = severity
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'B{row}'].fill = PatternFill(
            start_color=self.severity_colors.get(severity, "4CAF50"),
            end_color=self.severity_colors.get(severity, "4CAF50"),
            fill_type="solid"
        )
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Malicious:"
        is_malicious = risk.get('is_malicious', False)
        ws[f'B{row}'] = "YES" if is_malicious else "NO"
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'B{row}'].fill = PatternFill(
            start_color="F44336" if is_malicious else "4CAF50",
            end_color="F44336" if is_malicious else "4CAF50",
            fill_type="solid"
        )
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 2
        ws[f'A{row}'] = "THREAT INDICATORS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "Abuse Confidence:"
        ws[f'B{row}'] = f"{threat.get('abuse_confidence_score', 0)}%"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Fraud Score:"
        ws[f'B{row}'] = f"{threat.get('fraud_score', 0)}%"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Abuse Reports:"
        ws[f'B{row}'] = threat.get('abuse_reports', 0)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Blacklisted:"
        ws[f'B{row}'] = "YES" if threat.get('is_blacklisted', False) else "NO"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 2
        ws[f'A{row}'] = "LOCATION"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "Country:"
        ws[f'B{row}'] = geo.get('country', 'Unknown')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "City:"
        ws[f'B{row}'] = geo.get('city', 'Unknown')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "ISP:"
        ws[f'B{row}'] = geo.get('isp', 'Unknown')
        ws[f'A{row}'].font = Font(bold=True)
        
        # Recommendation
        row += 2
        ws[f'A{row}'] = "RECOMMENDATION"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = risk.get('recommendation', '')
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_network_location_sheet(self, wb: openpyxl.Workbook, scan_data: Dict[str, Any]):
        """Create Network & Location details sheet"""
        ws = wb.create_sheet("Network & Location")
        
        net = scan_data.get("network_info", {})
        geo = scan_data.get("geographic_details", {})
        threat = scan_data.get("threat_intelligence", {})
        
        # --- Network Information ---
        ws['A1'] = "Network Information"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')
        
        row = 3
        network_data = [
            ("ISP Name", net.get('isp', 'Unknown')),
            ("Organization", net.get('organization', 'Unknown')),
            ("ASN", net.get('asn', 'N/A')),
            ("ASN Name", net.get('asn_name', 'N/A')),
            ("Usage Type", threat.get('usage_type', 'General')),
            ("Connection Type", threat.get('connection_type', 'Unknown')),
        ]
        
        for label, value in network_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        # --- Geographic Details ---
        row += 2
        ws[f'A{row}'] = "Geographic Details"
        ws[f'A{row}'].font = Font(size=16, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 2
        geo_data = [
            ("Country", geo.get('country', 'Unknown')),
            ("City", geo.get('city', 'Unknown')),
            ("Region", geo.get('region', 'Unknown')),
            ("Timezone", geo.get('timezone', 'Unknown')),
            ("Latitude", geo.get('latitude', 0)),
            ("Longitude", geo.get('longitude', 0)),
            ("Postal Code", geo.get('postal', 'N/A')),
        ]
        
        for label, value in geo_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _create_threat_intel_sheet(self, wb: openpyxl.Workbook, scan_data: Dict[str, Any]):
        """Create enhanced threat user sheet"""
        ws = wb.create_sheet("Threat Intelligence")
        threat = scan_data.get("threat_intelligence", {})
        abuse = scan_data.get("abuse_history", {})
        
        ws['A1'] = "Threat Intelligence Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Indicators
        row = 3
        headers = ["Indicator", "Status", "Details"]
        for col, header in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=header)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            
        indicators = [
            ("VPN Detected", threat.get('is_vpn', False), "Using VPN service"),
            ("Proxy Detected", threat.get('is_proxy', False), "Using Proxy service"),
            ("Tor Exit Node", threat.get('is_tor_exit', False), "Tor Exit Node detected"),
            ("Blacklisted", threat.get('is_blacklisted', False), "Flagged on blacklists"),
            ("Bot Activity", threat.get('is_bot', False), "Automated bot behavior"),
        ]
        
        for ind, status, det in indicators:
            row += 1
            ws.cell(row=row, column=1, value=ind)
            s_cell = ws.cell(row=row, column=2, value="YES" if status else "NO")
            s_cell.font = Font(bold=True, color="FFFFFF")
            s_cell.fill = PatternFill(start_color="F44336" if status else "4CAF50", end_color="F44336" if status else "4CAF50", fill_type="solid")
            ws.cell(row=row, column=3, value=det)
            
        # Abuse History
        row += 3
        ws[f'A{row}'] = "Abuse History"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1
        abuse_data = [
            ("Abuse Confidence", f"{threat.get('abuse_confidence_score', 0)}%"),
            ("Total Reports", abuse.get('total_reports', 0)),
            ("Distinct Users", abuse.get('distinct_users', 0)),
            ("Last Reported", abuse.get('last_reported_at', 'Never')),
        ]
        for l, v in abuse_data:
            ws.cell(row=row, column=1, value=l).font = Font(bold=True)
            ws.cell(row=row, column=2, value=v)
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40

    def export_to_pdf(self, scan_data: Dict[str, Any]) -> BytesIO:
        """Export scan results to PDF with professional formatting"""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#1976D2'), alignment=TA_CENTER)
        h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=16, textColor=colors.HexColor('#1976D2'), spaceBefore=12)
        
        story.append(Paragraph("IP Security Scan Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # -- Assessment Table (More compact) --
        risk = scan_data.get("risk_assessment", {})
        severity = risk.get('severity', 'Low')
        sev_color = {"Low": colors.green, "Medium": colors.orange, "High": colors.orangered, "Critical": colors.red}.get(severity, colors.grey)
        
        risk_data = [
            ["Risk Score", f"{risk.get('score', 0)}/100"],
            ["Severity", severity],
            ["Status", "MALICIOUS" if risk.get('is_malicious') else "Clean"]
        ]
        t = Table(risk_data, colWidths=[2*inch, 4*inch])
        t.setStyle(TableStyle([
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('BACKGROUND', (1,1), (1,1), sev_color),
            ('TEXTCOLOR', (1,1), (1,1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('PADDING', (0,0), (-1,-1), 6)
        ]))
        story.append(t)
        story.append(Spacer(1, 0.2*inch))
        
        # -- Expanded Network & Location --
        story.append(Paragraph("Detailed Network & Location Analysis", h2_style))
        net = scan_data.get("network_info", {})
        geo = scan_data.get("geographic_details", {})
        
        net_data = [
            # Network Section
            ["Network Information", "", "", ""],
            ["ISP", net.get('isp', 'Unknown'), "ASN", net.get('asn', 'N/A')],
            ["Organization", net.get('organization', 'Unknown'), "ASN Name", net.get('asn_name', 'N/A')],
            ["Domain", net.get('domain', 'N/A'), "Net Range", net.get('network_range', 'N/A')],
            # Location Section
            ["Geographic Details", "", "", ""],
            ["Country", f"{geo.get('country', 'Unknown')} ({geo.get('country_code', '')})", "Continent", geo.get('continent', 'Unknown')],
            ["City", geo.get('city', 'Unknown'), "Region", f"{geo.get('region', '')} ({geo.get('region_name', '')})"],
            ["Zip Code", geo.get('zip_code', 'N/A'), "Timezone", geo.get('timezone', 'Unknown')],
            ["Coordinates", f"{geo.get('latitude',0)}, {geo.get('longitude',0)}", "Local Time", datetime.utcnow().strftime('%H:%M UTC')]
        ]
        
        net_table = Table(net_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
        net_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            # Section Headers
            ('SPAN', (0,0), (-1,0)), # Network Header span
            ('SPAN', (0,4), (-1,4)), # Location Header span
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E3F2FD')),
            ('BACKGROUND', (0,4), (-1,4), colors.HexColor('#E3F2FD')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME', (0,4), (-1,4), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('ALIGN', (0,4), (-1,4), 'CENTER'),
            # Field Labels
            ('FONTNAME', (0,1), (0,3), 'Helvetica-Bold'),
            ('FONTNAME', (2,1), (2,3), 'Helvetica-Bold'),
            ('FONTNAME', (0,5), (0,8), 'Helvetica-Bold'),
            ('FONTNAME', (2,5), (2,8), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(net_table)
        story.append(Spacer(1, 0.2*inch))
        
        # -- Expanded Threat Intel --
        story.append(Paragraph("Advanced Threat Intelligence", h2_style))
        threat = scan_data.get("threat_intelligence", {})
        abuse = scan_data.get("abuse_history", {})
        
        th_data = [
            ["Risk Indicator", "Status / Value", "Risk Indicator", "Status / Value"],
            # Row 1
            ["Abuse Confidence", f"{threat.get('abuse_confidence_score', 0)}%", "Fraud Score", f"{threat.get('fraud_score', 0)}%"],
            # Row 2
            ["Is Proxy?", "YES" if threat.get('is_proxy') else "NO", "Is VPN?", "YES" if threat.get('is_vpn') else "NO"],
            # Row 3
            ["Is Tor Exit?", "YES" if threat.get('is_tor_exit') else "NO", "Is Bot?", "YES" if threat.get('is_bot') else "NO"],
            # Row 4
            ["Is Datacenter?", "YES" if threat.get('is_datacenter') else "NO", "Is Mobile?", "YES" if threat.get('is_mobile') else "NO"],
            # Row 5
            ["Blacklisted?", "YES" if threat.get('is_blacklisted') else "NO", "Whitelisted?", "YES" if abuse.get('is_whitelisted') else "NO"],
            # Row 6
            ["Recent Reports", f"{threat.get('abuse_reports', 0)}", "Distinct Users", f"{abuse.get('distinct_users', 0)}"],
            # Row 7
            ["Usage Type", threat.get('usage_type', 'N/A'), "Last Reported", threat.get('last_reported') or 'Never']
        ]
        
        th_table = Table(th_data, colWidths=[1.75*inch, 1.75*inch, 1.75*inch, 1.75*inch])
        th_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BACKGROUND', (0,1), (0,-1), colors.whitesmoke), # Alternating cols for labels
            ('BACKGROUND', (2,1), (2,-1), colors.whitesmoke),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ]))
        story.append(th_table)
        
        # Risk Factors (Existing code handles this well, but let's ensure it's kept)
        factors = risk.get('factors', [])
        if factors:
            story.append(Paragraph("Verified Risk Factors", h2_style))
            f_data = [["Risk Factor", "Impact", "Description"]] + [[f['name'], f"+{f['impact']}", f['description']] for f in factors]
            f_table = Table(f_data, colWidths=[2*inch, 0.8*inch, 4.2*inch])
            f_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D32F2F')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            story.append(f_table)

        doc.build(story)
        output.seek(0)
        return output


# Singleton instance
export_service = ExportService()
